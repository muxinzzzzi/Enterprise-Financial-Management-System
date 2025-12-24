from __future__ import annotations

import base64
import json
import logging
import os
from typing import Any, Dict, List
from decimal import Decimal
from pathlib import Path

from flask import Flask, jsonify, redirect, render_template, request, send_file, session, url_for
from flask_cors import CORS

from config import DATA_DIR, get_settings
from database import init_db
from llm_client import LLMClient
from models.schemas import (
    AnalyticsQueryRequest,
    DocumentPayload,
    FeedbackRequest,
    KnowledgeRulePayload,
    PipelineOptions,
    PolicyDocument,
    ReconciliationRequest,
    ReviewUpdateRequest,
)
from pipelines.reconciliation_pipeline import ReconciliationPipeline
from repositories.audit_log import AuditLogger
from repositories.vector_store import VectorStore
from services.analytics.analytics_service import AnalyticsService
from services.analytics.anomaly_service import AnomalyDetectionService
from services.analytics.dashboard_service import summary as dashboard_summary
from services.analytics.feedback_service import FeedbackService
from services.assistants.assistant_service import AssistantService
from services.assistants.review_service import ReviewService
from services.extraction.categorization_service import ExpenseCategorizationService
from services.extraction.extraction_service import FieldExtractionService
from services.extraction.normalization_service import NormalizationService
from services.ingestion.ingestion_service import DocumentIngestionService
from services.ingestion.ocr_service import MultiEngineOCRService
from services.accounting.journal_service import JournalEntryService
from services.policy_rag.policy_service import PolicyValidationService
from services.accounting.persistence_service import (
    persist_results,
    _generate_voucher_pdf,
    generate_combined_voucher,
)
from services.analytics.report_service import ReportService
from services.analytics.advanced_report_service import AdvancedReportService
from services.accounting.ai_accountant import FinancialReportService as LegacyFinancialReportService
from services.financial_reports.report_service import FinancialReportService
from models.financial_schemas import ReportConfig
from services.policy_rag.knowledge_base_service import KnowledgeBaseService
from services.qa_service import QAService
from database import db_session
from models.db_models import Document, LedgerEntry
from models.schemas import DocumentResult, PolicyFlag
from services.user_service import (
    authenticate,
    create_user,
    get_user,
    list_users,
    login_or_register,
)
from utils.file_ops import save_base64_file
from openpyxl import Workbook, load_workbook
from datetime import datetime


logging.basicConfig(level=logging.INFO, format="[%(asctime)s] %(levelname)s %(message)s")

app = Flask(__name__)
CORS(app)
settings = get_settings()
init_db()
app.secret_key = os.getenv("APP_SECRET", "dev-secret")

llm_client = LLMClient(settings.llm_model, settings.llm_api_key, settings.llm_base_url)
vector_store = VectorStore()
journal_service = JournalEntryService(llm_client)
assistant_service = AssistantService(llm_client)
normalization_service = NormalizationService(settings=settings)
analytics_service = AnalyticsService(llm_client, cache_limit=settings.analytics_cache_limit)
legacy_financial_report_service = LegacyFinancialReportService()
financial_report_service = FinancialReportService(llm_client=llm_client)

services = {
    "ingestion": DocumentIngestionService(),
    "ocr": MultiEngineOCRService(llm_client),
    "extraction": FieldExtractionService(llm_client),
    "normalization": normalization_service,
    "policy": PolicyValidationService(vector_store, llm_client),
    "categorization": ExpenseCategorizationService(llm_client),
    "anomaly": AnomalyDetectionService(),
    "analytics": analytics_service,
    "feedback": FeedbackService(),
    "report": ReportService(llm_client),
}

# 初始化高级报表服务
advanced_report_service = AdvancedReportService(
    llm_client, output_dir=DATA_DIR / "reports"
)

knowledge_service = KnowledgeBaseService(llm_client, services["policy"])
review_service = ReviewService(services["feedback"], advanced_report_service, llm_client, policy_service=services["policy"])

try:
    # 如需从 shadow_rules.json 首次导入规则，可调用 seed_shadow_rules()
    knowledge_service.refresh_vector_store()
except Exception:
    logging.exception("init knowledge base refresh failed")

pipeline = ReconciliationPipeline(**services)
audit_logger = AuditLogger(DATA_DIR / "cache" / "audit.log")
app_state: dict[str, Any] = {"last_result": None}


@app.get("/")
def root() -> Any:
    if session.get("user_id"):
        return redirect(url_for("dashboard_page"))
    return redirect(url_for("login_page"))


@app.get("/login")
def login_page() -> Any:
    if session.get("user_id"):
        return redirect(url_for("dashboard_page"))
    return render_template("login.html")


@app.get("/dashboard")
def dashboard_page() -> Any:
    if not session.get("user_id"):
        return redirect(url_for("login_page"))
    return render_template("dashboard.html")


@app.get("/health")
def health() -> Any:
    return jsonify(
        {
            "status": "ok",
            "env": settings.env,
            "llm_enabled": bool(settings.llm_api_key),
            "data_dir": str(DATA_DIR),
        }
    )


@app.post("/api/v1/reconciliations/run")
def run_pipeline() -> Any:
    data = request.get_json(force=True)
    payload = ReconciliationRequest(**data)
    result = pipeline.run(payload)
    for doc in result.documents:
        doc.journal_entries = journal_service.generate(doc)
    app_state["last_result"] = result
    audit_logger.log("pipeline_run", {"job_id": result.job_id, "doc_count": len(result.documents)})
    persist_results(result.documents, payload.documents)
    return jsonify(result.model_dump())


@app.post("/api/v1/reconciliations/upload")
def upload_and_run() -> Any:
    if "file" not in request.files:
        return jsonify({"success": False, "error": "缺少 file 字段"}), 400

    file_storage = request.files["file"]
    if not file_storage or not file_storage.filename:
        return jsonify({"success": False, "error": "文件无效"}), 400

    meta = _parse_form_json(request.form.get("meta"), default={})
    options_payload = _parse_form_json(request.form.get("options"), default={})
    policies_payload = _parse_form_json(request.form.get("policies"), default=[])

    raw_bytes = file_storage.read()
    if not raw_bytes:
        return jsonify({"success": False, "error": "文件内容为空"}), 400

    base64_content = base64.b64encode(raw_bytes).decode("utf-8")
    file_path = save_base64_file(file_storage.filename, base64_content, sub_dir="input/captures")
    meta["file_path"] = str(file_path)

    document_payload = DocumentPayload(
        file_name=file_storage.filename,
        file_content_base64=base64_content,
        meta=meta,
    )
    user_id_value = request.form.get("user_id")
    request_model = ReconciliationRequest(
        documents=[document_payload],
        policies=[PolicyDocument(**policy) for policy in policies_payload],
        options=PipelineOptions(**options_payload),
        user_id=int(user_id_value) if user_id_value else None,
    )

    result = pipeline.run(request_model)
    for doc in result.documents:
        doc.journal_entries = journal_service.generate(doc)
    app_state["last_result"] = result
    audit_logger.log(
        "upload_pipeline_run",
        {
            "job_id": result.job_id,
            "doc_count": len(result.documents),
            "file_name": file_storage.filename,
        },
    )
    persist_results(result.documents, request_model.documents)
    return jsonify(
        {
            "success": True,
            "job_id": result.job_id,
            "document": result.documents[0].model_dump() if result.documents else {},
            "report": result.report,
        }
    )


@app.get("/api/v1/invoices")
def list_invoices() -> Any:
    user_id = request.args.get("user_id")
    q = request.args.get("q", "")
    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")
    page = max(1, int(request.args.get("page", 1) or 1))
    page_size = max(1, int(request.args.get("page_size", 6) or 6))
    with db_session() as session:
        query = session.query(Document)
        # 时间区间过滤（基于创建时间）
        if start_date:
            try:
                from datetime import datetime

                sd = datetime.fromisoformat(start_date)
                query = query.filter(Document.created_at >= sd)
            except Exception:
                pass
        if end_date:
            try:
                from datetime import datetime

                ed = datetime.fromisoformat(end_date)
                query = query.filter(Document.created_at <= ed)
            except Exception:
                pass
        if user_id:
            query = query.filter(Document.user_id == user_id)
        # 简单关键字搜索：文件名 / vendor / category / raw_result JSON 包含
        if q:
            like_q = f"%{q}%"
            query = query.filter(
                (Document.file_name.ilike(like_q))
                | (Document.vendor.ilike(like_q))
                | (Document.category.ilike(like_q))
            )
        total = query.count()

        # 先获取按时间升序的 id，用于生成稳定的数字 display_id（1 开始，最早=1）
        asc_ids = [doc.id for doc in query.order_by(Document.created_at.asc()).all()]
        id_to_display = {doc_id: idx + 1 for idx, doc_id in enumerate(asc_ids)}

        query = query.order_by(Document.created_at.desc()).offset((page - 1) * page_size).limit(page_size)
        docs = query.all()
        result = []
        for doc in docs:
            entries = [
                {
                    "id": le.id,
                    "debit_account": le.debit_account,
                    "credit_account": le.credit_account,
                    "amount": le.amount,
                    "memo": le.memo,
                    "created_at": le.created_at.isoformat(),
                }
                for le in doc.ledger_entries
            ]
            raw = doc.raw_result or {}
            # 尝试提取发票自身的开票/日期字段（可能在 raw 的不同位置）
            issue_date = None
            if isinstance(raw, dict):
                issue_date = raw.get("issue_date") or raw.get("normalized_fields", {}).get("issue_date")
                # 有时 issue_date 在 structured_fields 或 top-level 字段
                if not issue_date:
                    issue_date = raw.get("structured_fields", {}).get("issue_date")
            voucher_path = raw.get("voucher_pdf_path")
            voucher_url = url_for("get_voucher_pdf", doc_id=doc.id) if voucher_path else None
            result.append(
                {
                    "display_id": id_to_display.get(doc.id, doc.id),
                    "id": doc.id,
                    "file_name": doc.file_name,
                    "vendor": doc.vendor,
                    "issue_date": issue_date,
                    "amount": doc.amount,
                    "tax_amount": doc.tax_amount,
                    "currency": doc.currency,
                    "category": doc.category,
                    "status": doc.status,
                    "created_at": doc.created_at.isoformat(),
                    "voucher_pdf_url": voucher_url,
                    "entries": entries,
                }
            )
    return jsonify(
        {"success": True, "data": {"items": result, "total": total, "page": page, "page_size": page_size}}
    )



@app.get("/api/v1/invoices/<doc_id>/file")
def get_invoice_file(doc_id: str) -> Any:
    """返回原始上传的发票文件以供预览/下载。"""
    with db_session() as session:
        doc = session.query(Document).filter_by(id=doc_id).first()
        if not doc:
            return jsonify({"success": False, "error": "not_found"}), 404
        path = doc.file_path
        if not path or not os.path.exists(path):
            return jsonify({"success": False, "error": "file_not_found"}), 404
        # 使用 send_file 返回原始文件（不作为附件，便于直接在浏览器预览）
        return send_file(path, as_attachment=False, download_name=doc.file_name)



@app.delete("/api/v1/invoices/<doc_id>")
def delete_invoice(doc_id: str) -> Any:
    """删除指定发票记录及其生成的凭证与原始文件（如存在）。"""
    with db_session() as session:
        doc = session.query(Document).filter_by(id=doc_id).first()
        if not doc:
            return jsonify({"success": False, "error": "not_found"}), 404
        # 删除相关文件：原始文件与凭证 PDF（若存在）
        try:
            raw = doc.raw_result or {}
            file_path = doc.file_path
            voucher_path = raw.get("voucher_pdf_path")
            if file_path and os.path.exists(file_path):
                try:
                    os.remove(file_path)
                except Exception:
                    pass
            if voucher_path and os.path.exists(voucher_path):
                try:
                    os.remove(voucher_path)
                except Exception:
                    pass
        except Exception:
            pass
        # 删除数据库记录（ledger entries cascade 删除）
        session.delete(doc)
    return jsonify({"success": True})


@app.get("/api/v1/invoices/<doc_id>/voucher")
def get_voucher_pdf(doc_id: str) -> Any:
    with db_session() as session:
        doc = session.query(Document).filter_by(id=doc_id).first()
        if not doc:
            return jsonify({"success": False, "error": "not_found"}), 404
        raw = doc.raw_result or {}
        path = raw.get("voucher_pdf_path")
        if not path or not os.path.exists(path):
            return jsonify({"success": False, "error": "voucher_not_found"}), 404
        return send_file(path, as_attachment=True, download_name=f"{doc_id}.pdf")


@app.post("/api/v1/invoices/<doc_id>/voucher/generate")
def generate_voucher(doc_id: str) -> Any:
    """按需重新生成/补生成记账凭证 PDF。"""
    try:
        with db_session() as session:
            doc = session.query(Document).filter_by(id=doc_id).first()
            if not doc:
                return jsonify({"success": False, "error": "not_found"}), 404

            raw = doc.raw_result or {}
            # 构造 DocumentResult 用于已有的生成函数
            issue_date = (
                raw.get("issue_date")
                or raw.get("normalized_fields", {}).get("issue_date")
                or raw.get("structured_fields", {}).get("issue_date")
                or (doc.created_at.date().isoformat() if doc.created_at else None)
            )
            from models.schemas import DocumentResult

            doc_result = DocumentResult(
                document_id=doc.id,
                file_name=doc.file_name,
                vendor=doc.vendor,
                currency=doc.currency or "CNY",
                total_amount=doc.amount,
                tax_amount=doc.tax_amount,
                issue_date=issue_date,
                category=doc.category,
                structured_fields=raw.get("structured_fields") or {},
                normalized_fields=raw.get("normalized_fields") or {},
                ocr_confidence=raw.get("ocr_confidence") or 0.0,
                ocr_spans=raw.get("ocr_spans") or [],
                policy_flags=[],
                anomalies=[],
                duplicate_candidates=[],
                reasoning_trace=[],
                journal_entries=[],
            )
            entries = [
                {
                    "debit_account": le.debit_account,
                    "credit_account": le.credit_account,
                    "amount": le.amount,
                    "memo": le.memo,
                }
                for le in doc.ledger_entries
            ]

            voucher_path = _generate_voucher_pdf(doc_result, entries)
            if not voucher_path:
                return jsonify({"success": False, "error": "generate_failed"}), 500

            raw["voucher_pdf_path"] = str(voucher_path)
            # 直接 update JSON，避免部分环境下 JSON attr 未检测到变更
            session.query(Document).filter_by(id=doc.id).update({Document.raw_result: raw})

            voucher_url = url_for("get_voucher_pdf", doc_id=doc.id)
            return jsonify({"success": True, "voucher_pdf_url": voucher_url})
    except Exception as exc:
        logging.exception("generate voucher api failed: %s", doc_id)
        return jsonify({"success": False, "error": str(exc)}), 500


@app.post("/api/v1/invoices/batch_voucher/generate")
def batch_generate_vouchers() -> Any:
    """批量生成“合并凭证”：勾选多张发票 -> 生成一张凭证 PDF。"""
    data = request.get_json(force=True)
    invoice_ids = data.get("invoice_ids", [])
    if not invoice_ids or not isinstance(invoice_ids, list):
        return jsonify({"success": False, "error": "invoice_ids required"}), 400

    try:
        doc_results = []
        all_entries = []
        with db_session() as session:
            docs = session.query(Document).filter(Document.id.in_(invoice_ids)).all()
            if not docs:
                return jsonify({"success": False, "error": "未找到发票"}), 404

            for doc in docs:
                raw = doc.raw_result or {}
                issue_date = (
                    raw.get("issue_date")
                    or raw.get("normalized_fields", {}).get("issue_date")
                    or raw.get("structured_fields", {}).get("issue_date")
                    or (doc.created_at.date().isoformat() if doc.created_at else None)
                )
                doc_result = DocumentResult(
                    document_id=doc.id,
                    file_name=doc.file_name,
                    vendor=doc.vendor,
                    currency=doc.currency or "CNY",
                    total_amount=doc.amount,
                    tax_amount=doc.tax_amount,
                    issue_date=issue_date,
                    category=doc.category,
                    structured_fields=raw.get("structured_fields") or {},
                    normalized_fields=raw.get("normalized_fields") or {},
                    ocr_confidence=raw.get("ocr_confidence") or 0.0,
                    ocr_spans=raw.get("ocr_spans") or [],
                    policy_flags=[],
                    anomalies=[],
                    duplicate_candidates=[],
                    reasoning_trace=[],
                    journal_entries=[],
                )
                entries = journal_service.generate(doc_result)
                # 标记 memo，方便在 PDF 中区分来源
                for e in entries:
                    e["memo"] = e.get("memo") or doc.vendor or doc.file_name
                doc_results.append(doc_result)
                all_entries.extend(entries)

            voucher_path, voucher_no, voucher_date = generate_combined_voucher(doc_results, all_entries)
            if not voucher_path:
                return jsonify({"success": False, "error": "生成凭证失败"}), 500

            for doc in docs:
                raw = doc.raw_result or {}
                raw["voucher_pdf_path"] = str(voucher_path)
                raw["voucher_no"] = voucher_no
                raw["voucher_date"] = voucher_date
                session.query(Document).filter_by(id=doc.id).update(
                    {
                        Document.raw_result: raw,
                        Document.status: "voucher_generated",
                    }
                )

        return jsonify(
            {
                "success": True,
                "voucher_pdf_path": str(voucher_path),
                "voucher_no": voucher_no,
                "voucher_date": voucher_date,
                "invoice_count": len(invoice_ids),
            }
        )
    except Exception as exc:
        logging.exception("batch generate vouchers api failed")
        return jsonify({"success": False, "error": str(exc)}), 500


@app.get("/api/v1/vouchers")
def list_vouchers() -> Any:
    """获取凭证列表（合并凭证，按 voucher_pdf_path 聚合）。"""
    user_id = request.args.get("user_id")
    q = request.args.get("q", "")
    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")
    page = max(1, int(request.args.get("page", 1) or 1))
    page_size = max(1, int(request.args.get("page_size", 6) or 6))

    with db_session() as session:
        query = session.query(Document).filter(Document.raw_result.isnot(None))

        # 时间区间过滤
        if start_date:
            try:
                from datetime import datetime
                sd = datetime.fromisoformat(start_date)
                query = query.filter(Document.created_at >= sd)
            except Exception:
                pass
        if end_date:
            try:
                from datetime import datetime
                ed = datetime.fromisoformat(end_date)
                query = query.filter(Document.created_at <= ed)
            except Exception:
                pass
        if user_id:
            query = query.filter(Document.user_id == user_id)

        # 关键字搜索
        if q:
            like_q = f"%{q}%"
            query = query.filter(
                (Document.file_name.ilike(like_q))
                | (Document.vendor.ilike(like_q))
                | (Document.category.ilike(like_q))
            )

        query = query.order_by(Document.created_at.desc())
        docs = query.all()
        grouped: dict[str, dict] = {}

        for doc in docs:
            raw = doc.raw_result or {}
            voucher_path = raw.get("voucher_pdf_path")
            if not voucher_path:
                continue  # 只显示有凭证的

            key = voucher_path
            if key not in grouped:
                grouped[key] = {
                    "id": raw.get("voucher_no") or key,
                    "invoice_ids": [],
                    "total_amount": 0.0,
                    "created_at": raw.get("voucher_date") or (doc.created_at.isoformat() if doc.created_at else None),
                    "voucher_pdf_url": url_for("get_voucher_pdf", doc_id=doc.id),
                    "voucher_excel_url": None,
                    "invoices": [],
                    "voucher_no": raw.get("voucher_no"),
                    "voucher_date": raw.get("voucher_date"),
                }
            grouped[key]["invoice_ids"].append(doc.id)
            grouped[key]["total_amount"] += float(doc.amount or 0.0)
            grouped[key]["invoices"].append(
                {
                    "id": doc.id,
                    "file_name": doc.file_name,
                    "vendor": doc.vendor,
                    "amount": doc.amount,
                }
            )

        result = list(grouped.values())

        # 先按时间升序生成 display_id（最早=1），再按时间降序分页，最新在前
        def _created_at(row):
            return row.get("created_at") or ""

        for idx, row in enumerate(sorted(result, key=_created_at)):
            row["display_id"] = idx + 1

        result.sort(key=_created_at, reverse=True)
        total = len(result)
        start = (page - 1) * page_size
        end = start + page_size
        page_items = result[start:end]

    return jsonify(
        {"success": True, "data": {"items": page_items, "total": total, "page": page, "page_size": page_size}}
    )


@app.delete("/api/v1/vouchers")
def delete_vouchers() -> Any:
    """根据凭证号或包含的发票，删除已生成的凭证文件并解绑文档记录。"""
    payload = request.get_json(silent=True) or {}
    voucher_ids = payload.get("voucher_ids") or []
    invoice_ids = payload.get("invoice_ids") or []

    voucher_ids = [str(v).strip() for v in voucher_ids if str(v).strip()]
    invoice_ids = [str(i).strip() for i in invoice_ids if str(i).strip()]

    if not voucher_ids and not invoice_ids:
        return jsonify({"success": False, "error": "voucher_ids or invoice_ids required"}), 400

    removed_docs = 0
    removed_files: set[str] = set()

    with db_session() as session:
        docs = session.query(Document).filter(Document.raw_result.isnot(None)).all()
        for doc in docs:
            raw = doc.raw_result or {}
            voucher_path = raw.get("voucher_pdf_path")
            voucher_no = raw.get("voucher_no")
            voucher_excel_path = raw.get("voucher_excel_path")

            has_voucher = bool(voucher_path or voucher_no)
            if not has_voucher:
                continue

            path_stem = Path(voucher_path).stem if voucher_path else ""
            matched = False
            if voucher_ids:
                if voucher_no and str(voucher_no) in voucher_ids:
                    matched = True
                if path_stem and path_stem in voucher_ids:
                    matched = True
            if not matched and invoice_ids and doc.id in invoice_ids:
                matched = True

            if not matched:
                continue

            if voucher_excel_path:
                removed_files.add(str(voucher_excel_path))
            if voucher_path:
                removed_files.add(str(voucher_path))

            raw.pop("voucher_pdf_path", None)
            raw.pop("voucher_no", None)
            raw.pop("voucher_date", None)
            raw.pop("voucher_excel_path", None)
            session.query(Document).filter_by(id=doc.id).update(
                {
                    Document.raw_result: raw,
                    Document.status: "auto_recorded",
                }
            )
            removed_docs += 1

    for fp in removed_files:
        try:
            if fp and os.path.exists(fp):
                os.remove(fp)
        except Exception:
            pass

    return jsonify({"success": True, "removed": removed_docs, "files_removed": len(removed_files)})


@app.post("/api/v1/reports/financial")
def financial_report() -> Any:
    data = request.get_json(force=True)
    period = data.get("period_type", "month")  # month/quarter/year
    anchor = data.get("anchor_date")  # "2024-08-01"

    with db_session() as session:
        entries = session.query(LedgerEntry).all()
        rows = [
            {
                "account": f"{e.debit_account}/{e.credit_account}",
                "amount": e.amount,
                "memo": e.memo,
                "date": e.created_at.date().isoformat(),
            }
            for e in entries
        ]
    summary = legacy_financial_report_service.generate_summary(rows, period=period, anchor_date=anchor)
    narrative = legacy_financial_report_service.narrative(summary, llm_client)
    return jsonify({"success": True, "summary": summary, "report": narrative})



@app.post("/api/v1/invoices/voucher_excel")
def generate_voucher_excel() -> Any:
    """根据选中的发票列表生成 Excel 格式的记账凭证并返回文件。请求体: { invoice_ids: [id1, id2, ...] }"""
    data = request.get_json(force=True)
    invoice_ids = data.get("invoice_ids") or []
    if not invoice_ids or not isinstance(invoice_ids, list):
        return jsonify({"success": False, "error": "invoice_ids required"}), 400

    try:
        rows = []

        with db_session() as session:
            docs = session.query(Document).filter(Document.id.in_(invoice_ids)).all()
            for doc in docs:
                summary = doc.vendor or doc.file_name or ''
                # 为每个 ledger entry 生成借/贷行
                for le in doc.ledger_entries:
                    amt = float(le.amount or 0.0)
                    # 借方行
                    rows.append({"summary": summary, "account": le.debit_account or '', "debit": amt, "credit": None})
                    # 贷方行
                    rows.append({"summary": '', "account": le.credit_account or '', "debit": None, "credit": amt})

        # helper: convert amount to digit cells (string per cell), width 12
        def amount_to_cells(amount):
            if amount is None:
                return [""] * 12
            # amount as cents integer
            amt = Decimal(str(amount))
            cents = int((amt * 100).to_integral_value())
            s = str(abs(cents)).rjust(12, '0')  # pad to 12
            return list(s)

        # try to load template if exists
        template_paths = [
            DATA_DIR / 'templates' / 'voucher_template.xlsx',
            Path('voucher_template.xlsx'),
        ]
        wb = None
        ws = None
        for p in template_paths:
            try:
                if p.exists():
                    wb = load_workbook(p)
                    ws = wb.active
                    break
            except Exception:
                wb = None
                ws = None
        if wb is None:
            wb = Workbook()
            ws = wb.active

        start_row = 8
        for i, r in enumerate(rows):
            row = start_row + i
            ws.cell(row=row, column=1, value=r.get('summary', ''))  # A
            ws.cell(row=row, column=3, value=r.get('account', ''))  # C
            debit_cells = amount_to_cells(r.get('debit'))
            for k, v in enumerate(debit_cells):
                ws.cell(row=row, column=6 + k, value=v)  # F+
            credit_cells = amount_to_cells(r.get('credit'))
            for k, v in enumerate(credit_cells):
                ws.cell(row=row, column=18 + k, value=v)  # R+

        out_dir = DATA_DIR / 'output' / 'vouchers'
        out_dir.mkdir(parents=True, exist_ok=True)
        filename = out_dir / f"vouchers_{datetime.utcnow().strftime('%Y%m%d%H%M%S')}.xlsx"
        wb.save(filename)

        return send_file(str(filename), as_attachment=True, download_name=filename.name)
    except Exception as e:
        logging.exception("Failed to generate voucher excel")
        return jsonify({"success": False, "error": str(e)}), 500


@app.post("/api/v1/analytics/query")
def run_analytics() -> Any:
    data = request.get_json(force=True)
    request_model = AnalyticsQueryRequest(**data)
    insight = services["analytics"].query(request_model)
    return jsonify(insight.model_dump())


@app.post("/api/v1/auth/login")
def api_login() -> Any:
    data = request.get_json(force=True)
    email = data.get("email")
    password = data.get("password")
    if not email or not password:
        return jsonify({"success": False, "error": "邮箱和密码必填"}), 400
    user = authenticate(email=email, password=password)
    if not user:
        return jsonify({"success": False, "error": "账号或密码错误"}), 401
    session["user_id"] = user.id
    return jsonify({"success": True, "user": {"id": user.id, "name": user.name, "email": user.email}})


@app.post("/api/v1/auth/register")
def api_register() -> Any:
    data = request.get_json(force=True)
    name = data.get("name")
    email = data.get("email")
    password = data.get("password")
    if not all([name, email, password]):
        return jsonify({"success": False, "error": "姓名/邮箱/密码不能为空"}), 400
    user = create_user(name=name, email=email, password=password)
    return jsonify({"success": True, "user": {"id": user.id, "name": user.name, "email": user.email}})


@app.post("/api/v1/auth/logout")
def api_logout() -> Any:
    session.pop("user_id", None)
    return jsonify({"success": True})


@app.get("/api/v1/auth/me")
def api_me() -> Any:
    user_id = session.get("user_id")
    if not user_id:
        return jsonify({"success": False})
    user = get_user(user_id)
    if not user:
        session.pop("user_id", None)
        return jsonify({"success": False})
    return jsonify({"success": True, "user": {"id": user.id, "name": user.name, "email": user.email}})


@app.get("/api/v1/users")
def list_all_users() -> Any:
    return jsonify({"success": True, "users": list_users()})


@app.get("/api/v1/dashboard/summary")
def dashboard_summary_api() -> Any:
    days = int(request.args.get("days", 14))
    user_id = request.args.get("user_id") or session.get("user_id")
    data = dashboard_summary(days=days, user_id=user_id)
    return jsonify({"success": True, "data": data})


@app.post("/api/v1/feedback")
def collect_feedback() -> Any:
    data = request.get_json(force=True)
    request_model = FeedbackRequest(**data)
    services["feedback"].record(request_model)
    audit_logger.log("feedback", {"count": len(request_model.corrections)})
    return jsonify({"success": True})


@app.get("/api/v1/review/queue")
def review_queue() -> Any:
    status = request.args.get("status") or "uploaded"
    if status == "all":
        status = None
    q = request.args.get("q") or ""
    limit = min(200, int(request.args.get("limit", 50) or 50))
    try:
        data = review_service.list_queue(status=status, q=q, limit=limit)
        return jsonify({"success": True, "data": data})
    except Exception as exc:
        logging.exception("review queue error")
        return jsonify({"success": False, "error": str(exc)}), 500


@app.get("/api/v1/review/<doc_id>")
def review_detail(doc_id: str) -> Any:
    try:
        data = review_service.detail(doc_id)
        return jsonify({"success": True, "data": data})
    except Exception as exc:
        logging.exception("review detail error")
        return jsonify({"success": False, "error": str(exc)}), 404


@app.post("/api/v1/review/<doc_id>/update")
def review_update(doc_id: str) -> Any:
    try:
        payload = request.get_json(force=True)
        req = ReviewUpdateRequest(**payload)
        data = review_service.apply_changes(doc_id, req)
        return jsonify({"success": True, "data": data})
    except Exception as exc:
        logging.exception("review update error")
        return jsonify({"success": False, "error": str(exc)}), 400


@app.post("/api/v1/review/<doc_id>/approve")
def review_approve(doc_id: str) -> Any:
    body = request.get_json(silent=True) or {}
    reviewer_id = body.get("reviewer_id")
    comment = body.get("comment")
    try:
        data = review_service.approve(doc_id, reviewer_id=reviewer_id, comment=comment)
        return jsonify({"success": True, "data": data})
    except Exception as exc:
        logging.exception("review approve error")
        return jsonify({"success": False, "error": str(exc)}), 400


@app.post("/api/v1/review/batch/approve")
def review_batch_approve() -> Any:
    body = request.get_json(force=True) or {}
    doc_ids = body.get("doc_ids") or []
    reviewer_id = body.get("reviewer_id")
    if not isinstance(doc_ids, list) or not doc_ids:
        return jsonify({"success": False, "error": "doc_ids required"}), 400
    try:
        count = review_service.batch_approve(doc_ids, reviewer_id=reviewer_id)
        return jsonify({"success": True, "approved": count})
    except Exception as exc:
        logging.exception("batch approve error")
        return jsonify({"success": False, "error": str(exc)}), 400


@app.get("/api/v1/review/<doc_id>/reports")
def review_reports(doc_id: str) -> Any:
    try:
        data = review_service.generate_reports(doc_id)
        return jsonify({"success": True, "data": data})
    except Exception as exc:
        logging.exception("review report error")
        return jsonify({"success": False, "error": str(exc)}), 400


@app.get("/api/v1/review/training_samples")
def review_training_samples() -> Any:
    limit = int(request.args.get("limit", 8) or 8)
    try:
        samples = review_service.training_samples(limit=limit)
        return jsonify({"success": True, "data": samples})
    except Exception as exc:
        return jsonify({"success": False, "error": str(exc)}), 400


@app.post("/api/v1/review/backfill_policy")
def review_backfill_policy() -> Any:
    """批量为未跑过 RAG 的票据补充 policy_flags，避免点击详情时耗时调用。"""
    limit = int(request.args.get("limit", 50) or 50)
    try:
        meta = review_service.backfill_policy(limit=limit)
        return jsonify({"success": True, "data": meta})
    except Exception as exc:
        logging.exception("review backfill error")
        return jsonify({"success": False, "error": str(exc)}), 500


@app.get("/api/v1/knowledge/rules")
def list_rules() -> Any:
    q = request.args.get("q")
    category = request.args.get("category")
    page = int(request.args.get("page", 1) or 1)
    page_size = int(request.args.get("page_size", 10) or 10)
    try:
        data = knowledge_service.list_rules(q=q, category=category, page=page, page_size=page_size)
        return jsonify({"success": True, "data": data})
    except Exception as exc:
        logging.exception("knowledge list error")
        return jsonify({"success": False, "error": str(exc)}), 400


@app.get("/api/v1/knowledge/rules/<rule_id>")
def rule_detail(rule_id: str) -> Any:
    data = knowledge_service.get_rule(rule_id)
    if not data:
        return jsonify({"success": False, "error": "not_found"}), 404
    return jsonify({"success": True, "data": data})


@app.get("/api/v1/knowledge/rules/<rule_id>/versions")
def rule_versions(rule_id: str) -> Any:
    try:
        versions = knowledge_service.list_versions(rule_id)
        return jsonify({"success": True, "data": versions})
    except Exception as exc:
        return jsonify({"success": False, "error": str(exc)}), 400


@app.post("/api/v1/knowledge/rules")
def create_rule() -> Any:
    try:
        payload = request.get_json(force=True)
        req = KnowledgeRulePayload(**payload)
        user_id = payload.get("user_id") or session.get("user_id")
        data = knowledge_service.create_rule(req, user_id=user_id)
        return jsonify({"success": True, "data": data})
    except Exception as exc:
        logging.exception("create rule error")
        return jsonify({"success": False, "error": str(exc)}), 400


@app.put("/api/v1/knowledge/rules/<rule_id>")
def update_rule(rule_id: str) -> Any:
    try:
        payload = request.get_json(force=True)
        req = KnowledgeRulePayload(**payload)
        user_id = payload.get("user_id") or session.get("user_id")
        data = knowledge_service.update_rule(rule_id, req, user_id=user_id)
        return jsonify({"success": True, "data": data})
    except Exception as exc:
        logging.exception("update rule error")
        return jsonify({"success": False, "error": str(exc)}), 400


@app.post("/api/v1/knowledge/rules/refresh")
def refresh_rules() -> Any:
    try:
        meta = knowledge_service.refresh_vector_store()
        return jsonify({"success": True, "data": meta})
    except Exception as exc:
        logging.exception("refresh vector store error")
        return jsonify({"success": False, "error": str(exc)}), 400


@app.delete("/api/v1/knowledge/rules")
def delete_rules() -> Any:
    payload = request.get_json(force=True) or {}
    ids = payload.get("ids") or []
    if not isinstance(ids, list) or not ids:
        return jsonify({"success": False, "error": "ids required"}), 400
    try:
        meta = knowledge_service.delete_rules(ids)
        return jsonify({"success": True, "data": meta})
    except Exception as exc:
        logging.exception("delete rules error")
        return jsonify({"success": False, "error": str(exc)}), 400


@app.post("/api/v1/policies")
def upload_policy() -> Any:
    data = request.get_json(force=True)
    policy = PolicyDocument(**data)
    vector_store.add_texts([policy.content], [policy.model_dump()])
    return jsonify({"success": True})


@app.post("/api/v1/assistant/query")
def assistant_query() -> Any:
    data = request.get_json(force=True)
    question = data.get("question", "").strip()
    days = int(data.get("days", 120))
    user_id = data.get("user_id") or session.get("user_id")
    if not question:
        return jsonify({"success": False, "error": "question 不能为空"}), 400
    result = assistant_service.query(question=question, user_id=user_id, days=days)
    return jsonify({"success": True, **result})


@app.post("/api/v1/reports/invoice_audit")
def generate_invoice_audit_report() -> Any:
    """生成单张票据审核报告。
    
    请求体: {
        "document_id": "doc_xxx",
        "save_file": true  // 可选，是否保存文件
    }
    """
    data = request.get_json(force=True)
    document_id = data.get("document_id")
    save_file = data.get("save_file", True)

    if not document_id:
        return jsonify({"success": False, "error": "document_id 必填"}), 400

    try:
        with db_session() as session:
            doc = session.query(Document).filter(Document.id == document_id).first()
            if not doc:
                return jsonify({"success": False, "error": "票据不存在"}), 404

            # 从raw_result中提取数据
            raw_result = doc.raw_result or {}
            
            # 确保数值字段不为 None
            amount = doc.amount if doc.amount is not None else 0.0
            tax_amount = doc.tax_amount if doc.tax_amount is not None else 0.0
            ocr_confidence = raw_result.get("ocr_confidence")
            if ocr_confidence is None or not isinstance(ocr_confidence, (int, float)):
                ocr_confidence = 0.0
            
            # 构建DocumentResult
            document_result = DocumentResult(
                document_id=doc.id,
                file_name=doc.file_name,
                vendor=doc.vendor,
                currency=doc.currency,
                total_amount=amount,
                tax_amount=tax_amount,
                issue_date=raw_result.get("issue_date") or raw_result.get("normalized_fields", {}).get("issue_date"),
                category=doc.category,
                structured_fields=raw_result.get("structured_fields", {}),
                normalized_fields=raw_result.get("normalized_fields", {}),
                ocr_confidence=ocr_confidence,
                policy_flags=[PolicyFlag(**flag) if isinstance(flag, dict) else flag 
                             for flag in raw_result.get("policy_flags", [])],
                anomalies=raw_result.get("anomalies", []),
                duplicate_candidates=raw_result.get("duplicate_candidates", []),
            )

            # 提取数据
            policy_flags = document_result.policy_flags
            anomalies = document_result.anomalies
            duplicate_candidates = document_result.duplicate_candidates

            # 生成报告
            report = advanced_report_service.generate_invoice_audit_report(
                document_result, policy_flags, anomalies, duplicate_candidates, save_file
            )

            return jsonify({"success": True, "report": report, "document_id": document_id})
    except Exception as e:
        logging.exception("生成单张票据审核报告失败")
        return jsonify({"success": False, "error": str(e)}), 500


@app.post("/api/v1/reports/period_summary")
def generate_period_summary_report() -> Any:
    """生成周期汇总报表。
    
    请求体: {
        "start_date": "2025-01-01",  // 可选
        "end_date": "2025-01-31",    // 可选
        "user_id": "usr_xxx",        // 可选
        "period_type": "月",          // 可选，默认"月"
        "period_label": "2025年10期",  // 可选
        "save_file": true            // 可选
    }
    """
    data = request.get_json(force=True) or {}
    start_date = data.get("start_date")
    end_date = data.get("end_date")
    user_id = data.get("user_id") or session.get("user_id")
    period_type = data.get("period_type", "月")
    period_label = data.get("period_label", "")
    save_file = data.get("save_file", True)

    try:
        from datetime import datetime

        with db_session() as db:
            query = db.query(Document)

            # 时间过滤
            if start_date:
                try:
                    sd = datetime.fromisoformat(start_date)
                    query = query.filter(Document.created_at >= sd)
                except Exception:
                    pass
            if end_date:
                try:
                    ed = datetime.fromisoformat(end_date)
                    query = query.filter(Document.created_at <= ed)
                except Exception:
                    pass

            # 用户过滤
            if user_id:
                query = query.filter(Document.user_id == user_id)

            # 添加调试信息：查询总数
            total_count = db.query(Document).count()
            filtered_count = query.count()
            
            docs = query.order_by(Document.created_at.desc()).all()
            
            logging.info(f"周期报表查询: 总记录数={total_count}, 过滤后记录数={filtered_count}, 时间范围={start_date}~{end_date}, 用户ID={user_id}")

            # 转换为DocumentResult并提取数据
            documents = []
            all_policy_flags: Dict[str, List[PolicyFlag]] = {}
            all_anomalies: Dict[str, List[str]] = {}
            all_duplicates: Dict[str, List[str]] = {}

            for doc in docs:
                raw_result = doc.raw_result or {}
                
                document_result = DocumentResult(
                    document_id=doc.id,
                    file_name=doc.file_name,
                    vendor=doc.vendor,
                    currency=doc.currency,
                    total_amount=doc.amount,
                    tax_amount=doc.tax_amount,
                    issue_date=raw_result.get("issue_date") or raw_result.get("normalized_fields", {}).get("issue_date"),
                    category=doc.category,
                    structured_fields=raw_result.get("structured_fields", {}),
                    normalized_fields=raw_result.get("normalized_fields", {}),
                    ocr_confidence=raw_result.get("ocr_confidence", 0.0),
                    policy_flags=[],
                    anomalies=[],
                    duplicate_candidates=[],
                )

                documents.append(document_result)
                
                # 提取policy_flags, anomalies, duplicates
                policy_flags = [
                    PolicyFlag(**flag) if isinstance(flag, dict) else flag
                    for flag in raw_result.get("policy_flags", [])
                ]
                all_policy_flags[doc.id] = policy_flags
                all_anomalies[doc.id] = raw_result.get("anomalies", [])
                all_duplicates[doc.id] = raw_result.get("duplicate_candidates", [])

            # 生成报告
            report = advanced_report_service.generate_period_summary_report(
                documents,
                all_policy_flags,
                all_anomalies,
                all_duplicates,
                period_type,
                period_label,
                save_file,
            )

            return jsonify({
                "success": True,
                "report": report,
                "document_count": len(documents),
                "period_type": period_type,
                "period_label": period_label,
            })
    except Exception as e:
        logging.exception("生成周期汇总报表失败")
        return jsonify({"success": False, "error": str(e)}), 500


@app.post("/api/v1/reports/audit_trail")
def generate_audit_trail_report() -> Any:
    """生成审计追溯与整改清单。
    
    请求体: {
        "start_date": "2025-01-01",  // 可选
        "end_date": "2025-01-31",    // 可选
        "user_id": "usr_xxx",        // 可选
        "save_file": true            // 可选
    }
    """
    data = request.get_json(force=True) or {}
    start_date = data.get("start_date")
    end_date = data.get("end_date")
    user_id = data.get("user_id") or session.get("user_id")
    save_file = data.get("save_file", True)

    try:
        from datetime import datetime

        with db_session() as session:
            query = session.query(Document)

            # 时间过滤
            if start_date:
                try:
                    sd = datetime.fromisoformat(start_date)
                    query = query.filter(Document.created_at >= sd)
                except Exception:
                    pass
            if end_date:
                try:
                    ed = datetime.fromisoformat(end_date)
                    query = query.filter(Document.created_at <= ed)
                except Exception:
                    pass

            # 用户过滤
            if user_id:
                query = query.filter(Document.user_id == user_id)

            docs = query.order_by(Document.created_at.desc()).all()

            # 转换为DocumentResult并提取数据
            documents = []
            all_policy_flags: Dict[str, List[PolicyFlag]] = {}
            all_anomalies: Dict[str, List[str]] = {}
            all_duplicates: Dict[str, List[str]] = {}

            for doc in docs:
                raw_result = doc.raw_result or {}
                
                document_result = DocumentResult(
                    document_id=doc.id,
                    file_name=doc.file_name,
                    vendor=doc.vendor,
                    currency=doc.currency,
                    total_amount=doc.amount,
                    tax_amount=doc.tax_amount,
                    issue_date=raw_result.get("issue_date") or raw_result.get("normalized_fields", {}).get("issue_date"),
                    category=doc.category,
                    structured_fields=raw_result.get("structured_fields", {}),
                    normalized_fields=raw_result.get("normalized_fields", {}),
                    ocr_confidence=raw_result.get("ocr_confidence", 0.0),
                    policy_flags=[],
                    anomalies=[],
                    duplicate_candidates=[],
                )

                documents.append(document_result)
                
                # 提取policy_flags, anomalies, duplicates
                policy_flags = [
                    PolicyFlag(**flag) if isinstance(flag, dict) else flag
                    for flag in raw_result.get("policy_flags", [])
                ]
                all_policy_flags[doc.id] = policy_flags
                all_anomalies[doc.id] = raw_result.get("anomalies", [])
                all_duplicates[doc.id] = raw_result.get("duplicate_candidates", [])

            # 生成报告
            report = advanced_report_service.generate_audit_trail_report(
                documents, all_policy_flags, all_anomalies, all_duplicates, save_file
            )

            return jsonify({
                "success": True,
                "report": report,
                "document_count": len(documents),
            })
    except Exception as e:
        logging.exception("生成审计追溯与整改清单失败")
        return jsonify({"success": False, "error": str(e)}), 500


@app.post("/api/v1/reports/all")
def generate_all_reports() -> Any:
    """生成所有三类报表。
    
    请求体: {
        "start_date": "2025-01-01",  // 可选
        "end_date": "2025-01-31",    // 可选
        "user_id": "usr_xxx",        // 可选
        "period_type": "月",          // 可选
        "period_label": "2025年10期",  // 可选
        "save_files": true           // 可选
    }
    """
    data = request.get_json(force=True) or {}
    start_date = data.get("start_date")
    end_date = data.get("end_date")
    user_id = data.get("user_id") or session.get("user_id")
    period_type = data.get("period_type", "月")
    period_label = data.get("period_label", "")
    save_files = data.get("save_files", True)

    try:
        from datetime import datetime

        with db_session() as db:
            query = db.query(Document)

            # 时间过滤
            if start_date:
                try:
                    sd = datetime.fromisoformat(start_date)
                    query = query.filter(Document.created_at >= sd)
                except Exception:
                    pass
            if end_date:
                try:
                    ed = datetime.fromisoformat(end_date)
                    query = query.filter(Document.created_at <= ed)
                except Exception:
                    pass

            # 用户过滤
            if user_id:
                query = query.filter(Document.user_id == user_id)

            # 添加调试信息：查询总数
            total_count = db.query(Document).count()
            filtered_count = query.count()
            
            docs = query.order_by(Document.created_at.desc()).all()
            
            logging.info(f"所有报表查询: 总记录数={total_count}, 过滤后记录数={filtered_count}, 时间范围={start_date}~{end_date}, 用户ID={user_id}")

            # 转换为DocumentResult并提取数据
            documents = []
            all_policy_flags: Dict[str, List[PolicyFlag]] = {}
            all_anomalies: Dict[str, List[str]] = {}
            all_duplicates: Dict[str, List[str]] = {}

            for doc in docs:
                raw_result = doc.raw_result or {}
                
                document_result = DocumentResult(
                    document_id=doc.id,
                    file_name=doc.file_name,
                    vendor=doc.vendor,
                    currency=doc.currency,
                    total_amount=doc.amount,
                    tax_amount=doc.tax_amount,
                    issue_date=raw_result.get("issue_date") or raw_result.get("normalized_fields", {}).get("issue_date"),
                    category=doc.category,
                    structured_fields=raw_result.get("structured_fields", {}),
                    normalized_fields=raw_result.get("normalized_fields", {}),
                    ocr_confidence=raw_result.get("ocr_confidence", 0.0),
                    policy_flags=[],
                    anomalies=[],
                    duplicate_candidates=[],
                )

                documents.append(document_result)
                
                # 提取policy_flags, anomalies, duplicates
                policy_flags = [
                    PolicyFlag(**flag) if isinstance(flag, dict) else flag
                    for flag in raw_result.get("policy_flags", [])
                ]
                all_policy_flags[doc.id] = policy_flags
                all_anomalies[doc.id] = raw_result.get("anomalies", [])
                all_duplicates[doc.id] = raw_result.get("duplicate_candidates", [])

            # 生成所有报告
            reports = advanced_report_service.generate_all_reports(
                documents,
                all_policy_flags,
                all_anomalies,
                all_duplicates,
                period_type,
                period_label,
                save_files,
            )

            return jsonify({
                "success": True,
                "reports": reports,
                "document_count": len(documents),
                "period_type": period_type,
                "period_label": period_label,
            })
    except Exception as e:
        logging.exception("生成所有报表失败")
        return jsonify({"success": False, "error": str(e)}), 500


@app.post("/api/v1/financial-reports/balance-sheet")
def generate_balance_sheet() -> Any:
    """生成资产负债表。
    
    请求体: {
        "start_date": "2025-01-01",  // 可选
        "end_date": "2025-01-31",    // 可选
        "user_id": "usr_xxx",        // 可选
        "period_type": "month",       // 可选：month/quarter/year
        "currency": "CNY",            // 可选
        "company_name": "公司名称",    // 可选
        "enable_ai_analysis": false   // 可选
    }
    """
    data = request.get_json(force=True) or {}
    user_id = data.get("user_id") or session.get("user_id")
    
    try:
        from datetime import datetime
        
        start_date = None
        end_date = None
        if data.get("start_date"):
            start_date = datetime.fromisoformat(data["start_date"])
        if data.get("end_date"):
            end_date = datetime.fromisoformat(data["end_date"])
        
        config = ReportConfig(
            start_date=start_date,
            end_date=end_date,
            period_type=data.get("period_type", "month"),
            currency=data.get("currency", "CNY"),
            user_id=user_id,
            enable_ai_analysis=data.get("enable_ai_analysis", False),
            company_name=data.get("company_name"),
        )
        
        result = financial_report_service.generate_balance_sheet(config)
        
        # 保存markdown文件（服务层已生成PDF，这里只保存markdown）
        markdown_path = None
        if result.markdown_content:
            try:
                reports_dir = DATA_DIR / "reports" / "financial"
                reports_dir.mkdir(parents=True, exist_ok=True)
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                
                # 保存markdown
                md_filename = f"balance_sheet_{timestamp}.md"
                markdown_path = reports_dir / md_filename
                markdown_path.write_text(result.markdown_content, encoding="utf-8")
                logging.info(f"资产负债表Markdown已保存: {markdown_path}")
            except Exception as e:
                logging.error(f"保存Markdown失败: {e}")
        
        # 使用服务层返回的PDF路径（服务层已生成PDF）
        pdf_path = result.pdf_path
        
        return jsonify({
            "success": True,
            "report_type": result.report_type,
            "report_data": result.report_data,
            "markdown_content": result.markdown_content,
            "markdown_path": str(markdown_path) if markdown_path else None,
            "pdf_path": pdf_path,
            "ai_analysis": result.ai_analysis,
            "generated_at": result.generated_at.isoformat(),
        })
    except Exception as e:
        logging.exception("生成资产负债表失败")
        return jsonify({"success": False, "error": str(e)}), 500


@app.post("/api/v1/financial-reports/income-statement")
def generate_income_statement() -> Any:
    """生成利润表。
    
    请求体: {
        "start_date": "2025-01-01",  // 可选
        "end_date": "2025-01-31",    // 可选
        "user_id": "usr_xxx",        // 可选
        "period_type": "month",       // 可选：month/quarter/year
        "currency": "CNY",            // 可选
        "company_name": "公司名称",    // 可选
        "enable_ai_analysis": false   // 可选
    }
    """
    data = request.get_json(force=True) or {}
    user_id = data.get("user_id") or session.get("user_id")
    
    try:
        from datetime import datetime
        
        start_date = None
        end_date = None
        if data.get("start_date"):
            start_date = datetime.fromisoformat(data["start_date"])
        if data.get("end_date"):
            end_date = datetime.fromisoformat(data["end_date"])
        
        config = ReportConfig(
            start_date=start_date,
            end_date=end_date,
            period_type=data.get("period_type", "month"),
            currency=data.get("currency", "CNY"),
            user_id=user_id,
            enable_ai_analysis=data.get("enable_ai_analysis", False),
            company_name=data.get("company_name"),
        )
        
        result = financial_report_service.generate_income_statement(config)
        
        # 保存markdown文件（服务层已生成PDF，这里只保存markdown）
        markdown_path = None
        if result.markdown_content:
            try:
                reports_dir = DATA_DIR / "reports" / "financial"
                reports_dir.mkdir(parents=True, exist_ok=True)
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                
                # 保存markdown
                md_filename = f"income_statement_{timestamp}.md"
                markdown_path = reports_dir / md_filename
                markdown_path.write_text(result.markdown_content, encoding="utf-8")
                logging.info(f"利润表Markdown已保存: {markdown_path}")
            except Exception as e:
                logging.error(f"保存Markdown失败: {e}")
        
        # 如果服务层没有生成PDF，在这里生成（利润表暂时使用markdown转换）
        pdf_path = result.pdf_path
        if not pdf_path and result.markdown_content:
            try:
                from services.financial_reports.exporters.pdf_exporter import PDFExporter
                pdf_exporter = PDFExporter()
                pdf_path = pdf_exporter.export_income_statement(result.markdown_content, config)
                if pdf_path:
                    logging.info(f"利润表PDF已生成: {pdf_path}")
            except Exception as e:
                logging.error(f"生成利润表PDF失败: {e}")
        
        return jsonify({
            "success": True,
            "report_type": result.report_type,
            "report_data": result.report_data,
            "markdown_content": result.markdown_content,
            "markdown_path": str(markdown_path) if markdown_path else None,
            "pdf_path": pdf_path,
            "ai_analysis": result.ai_analysis,
            "generated_at": result.generated_at.isoformat(),
        })
    except Exception as e:
        logging.exception("生成利润表失败")
        return jsonify({"success": False, "error": str(e)}), 500


@app.post("/api/v1/financial-reports/cash-flow")
def generate_cash_flow() -> Any:
    """生成现金流量表。
    
    请求体: {
        "start_date": "2025-01-01",  // 可选
        "end_date": "2025-01-31",    // 可选
        "user_id": "usr_xxx",        // 可选
        "period_type": "month",       // 可选：month/quarter/year
        "currency": "CNY",            // 可选
        "company_name": "公司名称",    // 可选
        "enable_ai_analysis": false   // 可选
    }
    """
    data = request.get_json(force=True) or {}
    user_id = data.get("user_id") or session.get("user_id")
    
    try:
        from datetime import datetime
        
        start_date = None
        end_date = None
        if data.get("start_date"):
            start_date = datetime.fromisoformat(data["start_date"])
        if data.get("end_date"):
            end_date = datetime.fromisoformat(data["end_date"])
        
        config = ReportConfig(
            start_date=start_date,
            end_date=end_date,
            period_type=data.get("period_type", "month"),
            currency=data.get("currency", "CNY"),
            user_id=user_id,
            enable_ai_analysis=data.get("enable_ai_analysis", False),
            company_name=data.get("company_name"),
        )
        
        result = financial_report_service.generate_cash_flow(config)
        
        # 保存markdown文件（服务层已生成PDF，这里只保存markdown）
        markdown_path = None
        if result.markdown_content:
            try:
                reports_dir = DATA_DIR / "reports" / "financial"
                reports_dir.mkdir(parents=True, exist_ok=True)
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                
                # 保存markdown
                md_filename = f"cash_flow_{timestamp}.md"
                markdown_path = reports_dir / md_filename
                markdown_path.write_text(result.markdown_content, encoding="utf-8")
                logging.info(f"现金流量表Markdown已保存: {markdown_path}")
            except Exception as e:
                logging.error(f"保存Markdown失败: {e}")
        
        # 如果服务层没有生成PDF，在这里生成（现金流量表暂时使用markdown转换）
        pdf_path = result.pdf_path
        if not pdf_path and result.markdown_content:
            try:
                from services.financial_reports.exporters.pdf_exporter import PDFExporter
                pdf_exporter = PDFExporter()
                pdf_path = pdf_exporter.export_cash_flow(result.markdown_content, config)
                if pdf_path:
                    logging.info(f"现金流量表PDF已生成: {pdf_path}")
            except Exception as e:
                logging.error(f"生成现金流量表PDF失败: {e}")
        
        return jsonify({
            "success": True,
            "report_type": result.report_type,
            "report_data": result.report_data,
            "markdown_content": result.markdown_content,
            "markdown_path": str(markdown_path) if markdown_path else None,
            "pdf_path": pdf_path,
            "ai_analysis": result.ai_analysis,
            "generated_at": result.generated_at.isoformat(),
        })
    except Exception as e:
        logging.exception("生成现金流量表失败")
        return jsonify({"success": False, "error": str(e)}), 500


@app.post("/api/v1/financial-reports/all")
def generate_all_financial_reports() -> Any:
    """生成所有三类财务报表。
    
    请求体: {
        "start_date": "2025-01-01",  // 可选
        "end_date": "2025-01-31",    // 可选
        "user_id": "usr_xxx",        // 可选
        "period_type": "month",       // 可选：month/quarter/year
        "currency": "CNY",            // 可选
        "company_name": "公司名称",    // 可选
        "enable_ai_analysis": false   // 可选
    }
    """
    data = request.get_json(force=True) or {}
    user_id = data.get("user_id") or session.get("user_id")
    
    try:
        from datetime import datetime
        
        start_date = None
        end_date = None
        if data.get("start_date"):
            start_date = datetime.fromisoformat(data["start_date"])
        if data.get("end_date"):
            end_date = datetime.fromisoformat(data["end_date"])
        
        config = ReportConfig(
            start_date=start_date,
            end_date=end_date,
            period_type=data.get("period_type", "month"),
            currency=data.get("currency", "CNY"),
            user_id=user_id,
            enable_ai_analysis=data.get("enable_ai_analysis", False),
            company_name=data.get("company_name"),
        )
        
        results = financial_report_service.generate_all_reports(config)
        
        # 保存所有报表的markdown文件（服务层已生成PDF，这里只保存markdown）
        saved_reports = {}
        for report_type, result in results.items():
            markdown_path = None
            if result.markdown_content:
                try:
                    reports_dir = DATA_DIR / "reports" / "financial"
                    reports_dir.mkdir(parents=True, exist_ok=True)
                    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                    
                    # 保存markdown
                    md_filename = f"{report_type}_{timestamp}.md"
                    markdown_path = reports_dir / md_filename
                    markdown_path.write_text(result.markdown_content, encoding="utf-8")
                    logging.info(f"{report_type} Markdown已保存: {markdown_path}")
                except Exception as e:
                    logging.error(f"保存{report_type} Markdown失败: {e}")
            
            # 使用服务层返回的PDF路径（服务层已生成PDF）
            pdf_path = result.pdf_path
            
            # 如果服务层没有生成PDF，尝试生成（仅对利润表和现金流量表）
            if not pdf_path and result.markdown_content:
                try:
                    from services.financial_reports.exporters.pdf_exporter import PDFExporter
                    pdf_exporter = PDFExporter()
                    if report_type == "income_statement":
                        pdf_path = pdf_exporter.export_income_statement(result.markdown_content, config)
                    elif report_type == "cash_flow":
                        pdf_path = pdf_exporter.export_cash_flow(result.markdown_content, config)
                    if pdf_path:
                        logging.info(f"{report_type} PDF已生成: {pdf_path}")
                except Exception as e:
                    logging.error(f"生成{report_type} PDF失败: {e}")
            
            saved_reports[report_type] = {
                "report_type": result.report_type,
                "report_data": result.report_data,
                "markdown_content": result.markdown_content,
                "markdown_path": str(markdown_path) if markdown_path else None,
                "pdf_path": pdf_path,
                "ai_analysis": result.ai_analysis,
                "generated_at": result.generated_at.isoformat(),
            }
        
        return jsonify({
            "success": True,
            "reports": saved_reports,
        })
    except Exception as e:
        logging.exception("生成所有财务报表失败")
        return jsonify({"success": False, "error": str(e)}), 500


@app.post("/api/v1/financial-reports/<report_type>/pdf")
def generate_pdf(report_type: str) -> Any:
    """生成PDF文件。
    
    report_type: balance-sheet / income-statement / cash-flow
    """
    data = request.get_json(force=True) or {}
    user_id = data.get("user_id") or session.get("user_id")
    
    try:
        from datetime import datetime
        
        start_date = None
        end_date = None
        if data.get("start_date"):
            start_date = datetime.fromisoformat(data["start_date"])
        if data.get("end_date"):
            end_date = datetime.fromisoformat(data["end_date"])
        
        config = ReportConfig(
            start_date=start_date,
            end_date=end_date,
            period_type=data.get("period_type", "month"),
            currency=data.get("currency", "CNY"),
            user_id=user_id,
            enable_ai_analysis=False,
            company_name=data.get("company_name"),
        )
        
        # 生成报表
        if report_type == "balance-sheet":
            result = financial_report_service.generate_balance_sheet(config)
        elif report_type == "income-statement":
            result = financial_report_service.generate_income_statement(config)
        elif report_type == "cash-flow":
            result = financial_report_service.generate_cash_flow(config)
        else:
            return jsonify({"success": False, "error": f"未知的报表类型: {report_type}"}), 400
        
        # 生成PDF
        pdf_path = None
        if result.markdown_content:
            try:
                from services.financial_reports.exporters.pdf_exporter import PDFExporter
                pdf_exporter = PDFExporter()
                
                if report_type == "balance-sheet":
                    from models.financial_schemas import BalanceSheetData
                    balance_sheet_data = BalanceSheetData(**result.report_data)
                    # 如果启用AI分析，先生成分析
                    ai_analysis = None
                    if config.enable_ai_analysis:
                        from services.financial_reports.ai_analyzer import AIAnalyzer
                        ai_analyzer = AIAnalyzer(llm_client=llm_client)
                        ai_analysis = ai_analyzer.analyze_balance_sheet(balance_sheet_data)
                    pdf_path = pdf_exporter.export_balance_sheet(
                        balance_sheet_data, 
                        config, 
                        ai_analysis=ai_analysis
                    )
                elif report_type == "income-statement":
                    pdf_path = pdf_exporter.export_income_statement(result.markdown_content, config)
                elif report_type == "cash-flow":
                    pdf_path = pdf_exporter.export_cash_flow(result.markdown_content, config)
                
                if pdf_path:
                    logging.info(f"{report_type} PDF已生成: {pdf_path}")
            except Exception as e:
                logging.error(f"生成PDF失败: {e}")
        
        return jsonify({
            "success": True,
            "pdf_path": pdf_path,
            "markdown_path": str(DATA_DIR / "reports" / "financial" / f"{report_type.replace('-', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.md") if result.markdown_content else None,
        })
    except Exception as e:
        logging.exception(f"生成{report_type} PDF失败")
        return jsonify({"success": False, "error": str(e)}), 500


@app.get("/api/v1/financial-reports/pdf/<path:file_path>")
def get_pdf(file_path: str) -> Any:
    """获取PDF文件（预览或下载）。"""
    try:
        from urllib.parse import unquote
        
        # 解码文件路径
        decoded_path = unquote(file_path)
        file_path_obj = Path(decoded_path)
        
        # 安全检查：确保文件在reports目录下
        reports_dir = DATA_DIR / "reports" / "financial"
        if not str(file_path_obj).startswith(str(reports_dir)):
            return jsonify({"success": False, "error": "无效的文件路径"}), 403
        
        # 如果文件不存在，返回404
        if not file_path_obj.exists():
            return jsonify({"success": False, "error": "文件不存在"}), 404
        
        # 检查是否是下载请求
        download = request.args.get("download", "0") == "1"
        
        # 如果是PDF文件，直接返回
        if file_path_obj.suffix.lower() == ".pdf":
            return send_file(
                str(file_path_obj),
                mimetype="application/pdf",
                as_attachment=download,
                download_name=file_path_obj.name if download else None,
            )
        
        # 如果是markdown文件，转换为PDF
        if file_path_obj.suffix.lower() == ".md":
            try:
                from services.financial_reports.exporters.pdf_exporter import PDFExporter
                from models.financial_schemas import ReportConfig
                
                # 读取markdown内容
                markdown_content = file_path_obj.read_text(encoding="utf-8")
                
                # 生成PDF
                pdf_exporter = PDFExporter()
                pdf_path = file_path_obj.with_suffix('.pdf')
                
                # 根据文件名判断报表类型
                filename = file_path_obj.stem
                if 'balance_sheet' in filename:
                    # 资产负债表使用markdown转换方法（因为这里只有markdown文件）
                    pdf_path_str = pdf_exporter.export_balance_sheet_from_markdown(markdown_content, ReportConfig())
                elif 'income_statement' in filename:
                    pdf_path_str = pdf_exporter.export_income_statement(markdown_content, ReportConfig())
                elif 'cash_flow' in filename:
                    pdf_path_str = pdf_exporter.export_cash_flow(markdown_content, ReportConfig())
                else:
                    # 通用转换
                    if pdf_exporter._markdown_to_pdf(markdown_content, pdf_path):
                        pdf_path_str = str(pdf_path)
                    else:
                        raise Exception("PDF生成失败")
                
                if pdf_path_str and Path(pdf_path_str).exists():
                    return send_file(
                        pdf_path_str,
                        mimetype="application/pdf",
                        as_attachment=download,
                        download_name=Path(pdf_path_str).name if download else None,
                    )
                else:
                    raise Exception("PDF文件不存在")
            except Exception as e:
                logging.error(f"Markdown转PDF失败: {e}")
                # 如果转换失败，返回markdown内容
                if download:
                    return send_file(
                        str(file_path_obj),
                        mimetype="text/markdown",
                        as_attachment=True,
                        download_name=file_path_obj.name,
                    )
                else:
                    content = file_path_obj.read_text(encoding="utf-8")
                    return jsonify({
                        "success": True,
                        "content": content,
                        "type": "markdown",
                        "error": str(e),
                    })
        
        return jsonify({"success": False, "error": "不支持的文件类型"}), 400
    except Exception as e:
        logging.exception("获取PDF文件失败")
        return jsonify({"success": False, "error": str(e)}), 500


@app.errorhandler(Exception)
def handle_exception(error: Exception):  # type: ignore[override]
    logging.exception("API Error")
    return jsonify({"success": False, "error": str(error)}), 500


def _parse_form_json(value: str | None, default):
    if not value:
        return default
    try:
        return json.loads(value)
    except json.JSONDecodeError:
        return default


@app.post("/api/v1/qa/ask")
def qa_ask() -> Any:
    """智能问答接口。
    
    Body: { question, start_date?, end_date?, limit? }
    Return: { answer_md, evidence:{sql, params, rows_preview}, followups? }
    """
    data = request.get_json(force=True) or {}
    
    question = data.get("question", "").strip()
    if not question:
        return jsonify({"success": False, "error": "问题不能为空"}), 400
    
    start_date = data.get("start_date")
    end_date = data.get("end_date")
    limit = data.get("limit")
    
    try:
        # 创建QA服务实例
        with db_session() as db:
            qa_service = QAService(db, llm_client)
            result = qa_service.ask(question, start_date, end_date, limit)
        
        return jsonify({
            "success": True,
            "answer_md": result["answer_md"],
            "evidence": result["evidence"],
            "followups": result["followups"],
        })
    except Exception as e:
        logging.exception("智能问答失败")
        return jsonify({
            "success": False,
            "error": str(e),
            "answer_md": f"抱歉，处理您的问题时发生错误：{str(e)}",
            "evidence": None,
            "followups": None,
        }), 500


if __name__ == "__main__":
    import os

    port = int(os.getenv("PORT", "9000"))
    app.run(host="0.0.0.0", port=port, debug=settings.debug)
